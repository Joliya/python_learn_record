# -*- coding: utf-8 -*-
"""
excel 数据处理，读写 excel 格式文件, 只适合小文件读取， 大文件请自行使用 pandas 库处理
"""

from __future__ import absolute_import

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import InvalidFileException


class ExcelDataTypes(object):
    """
    excel导入的数据类型
    """
    EMPTY = 0
    STRING = 1
    NUMBER = 2
    DATE = 3
    BOOL = 4
    ERROR = 5


class ExcelDateMode(object):
    """
    excel导入的日期转换成datetime是应用的模式
    """
    BASED_ON_1900 = 0
    BASED_ON_1904 = 1


def get_data_from_excel(file_path=None, start_row=0, column_indexes=None, sheet_index=0, file_contents=None,
                        trim_empty_rows=True):
    """
    从excel文件中读取单个sheet的数据
    :param file_path: excel文件路径
    :param start_row: 从第几行读取，从0开始
    :param column_indexes: 读取哪几列，从0开始
    :param sheet_index: 读取哪个sheet，从0开始
    :param file_contents: 二进制excel文件内容
    :param trim_empty_rows, 是否过滤掉空行
    :return list of list
    """
    sheet_data_list = get_full_data_from_excel(file_path, file_contents=file_contents,
                                               start_sheet_index=sheet_index, sheet_num=1,
                                               trim_empty_rows=trim_empty_rows)

    data = sheet_data_list[0]
    if start_row:
        data = data[start_row:]
    if not column_indexes:
        return data
    result = []
    for row in data:
        row = [row[index] for index in column_indexes]
        result.append(row)
    return result


def get_full_data_from_excel(file_path=None, file_contents=None, start_sheet_index=0, sheet_num=None,
                             trim_empty_rows=True):
    """
    从excel文件中读取所有的sheet数据
    :param file_path: excel文件路径
    :param file_contents: 二进制excel文件内容
    :param start_sheet_index, 要取的sheet index编号，从0开始
    :param sheet_num, 要取的sheet的数量
    :param trim_empty_rows, 是否过滤掉空行
    :return list[sheet[row list]], sheet的列表，每个元素为二位数组list(list)
    """
    result = []

    try:
        if file_path:
            work_book = load_workbook(file_path)
        else:
            work_book = load_workbook(BytesIO(file_contents))
    except (InvalidFileException, IOError) as e:
        raise Exception('excel文件读取失败，仅支持xlsx格式，请检查文件是否正确。%s' % e)

    sheet_num = sheet_num or len(work_book.worksheets)  # 如果未指定sheet_num, 取start_sheet_index后的所有sheet
    for sheet in work_book.worksheets[start_sheet_index:start_sheet_index + sheet_num]:
        # 读取excel
        sheet_data = []
        for row in sheet.rows:
            fields = [cell.value for cell in row]
            if not fields:
                continue
            if trim_empty_rows and set(fields) == {None}:
                continue
            sheet_data.append(fields)
        # 如果sheet没有数据，也放到数组，确保sheet的index不变化
        result.append(sheet_data)
    return result


def write_file_as_xls(file_name, list_of_list, *more_lists, **kwargs):
    """
    写入文件到xlsx，不支持xls，此函数只用于旧版本函数签名兼容
    """
    return write_file_as_xlsx(file_name, list_of_list, *more_lists, **kwargs)


def write_data_as_memory_xlsx(list_of_list, *more_lists, **kwargs):
    """
    将数据写入到内存excel文件中
    """
    sio = BytesIO()
    write_file_as_xlsx(sio, list_of_list, *more_lists, **kwargs)
    sio.seek(0)
    return sio.getvalue()


def write_file_as_xlsx(file_name, list_of_list, *more_lists, **kwargs):
    """
    将文件写入xlsx文件，参数：要写入文件的名字，要写入的文件列表
    :param file_name: 文件名或者可io seek/write的对象tmp/result.xlsx'
    :param list_of_list: [title, row1, row2, ...]; 其中 title, row1, row2的格式: [cell1, cell2, ...]
    :param more_lists: tuple of list of list
    :param kwargs:
        sheet_names: list of sheet_name，长度等于 1 + len(more_lists)
    :return:
    """
    if isinstance(file_name, str):
        assert file_name.endswith(".xlsx"), "file_name must end with .xlsx"

    def _write_sheet(book, name, content):
        """
        将内容写入一个sheet
        :param book: excel文件
        :param name: sheet名称
        :param content: list of list 写入sheet中的内容
        :return:
        """
        sheet = book.create_sheet(name)
        for row_index, row in enumerate(content):
            # 包含不可见的特殊字符(ILLEGAL_CHARACTERS_RE)的字符串会导致excel写入失败，在校验时抛出IllegalCharacterError。这里提前删了
            row = [ILLEGAL_CHARACTERS_RE.sub("", cell) if isinstance(cell, str) else cell for cell in row]
            sheet.append(row)

    workbook = Workbook(write_only=True)
    if list_of_list:
        sheet_content_list = [list_of_list] + list(more_lists)
    else:
        sheet_content_list = list(more_lists)
    sheet_names = kwargs.get('sheet_names')
    # 如果表名列表长度和实际不一样，不使用
    if not sheet_names or len(sheet_names) != len(sheet_content_list):
        sheet_names = ["sheet_%s" % i for i in range(len(sheet_content_list))]
    for sheet_index, sheet_content in enumerate(sheet_content_list):
        _write_sheet(book=workbook, name=sheet_names[sheet_index], content=sheet_content)

    workbook.save(file_name)


def export_to_txt_file(filename, rows, field_sep='\t', line_sep='\n'):
    """
    导出到文件
    """
    with open(filename, 'w') as f:
        result = []
        for row in rows:
            row = ['%s' % item for item in row]
            result.append(field_sep.join(row))
        f.write(line_sep.join(result))
